import configparser
import json
import logging
import os
import spotipy
import threading
from openpyxl import load_workbook
from flask import Flask, redirect, request, session, render_template
from flask_session import Session

threads = {}
config = configparser.ConfigParser()
config.read('config')

client_id = config.get(section = 'Spotify', option = 'client_id', fallback = os.getenv('CLIENT_ID', default=None))
client_secret = config.get(section = 'Spotify', option = 'client_secret', fallback = os.getenv('CLIENT_SECRET', default=None))
redirect_uri = config.get(section = 'Spotify', option = 'redirect_uri', fallback = os.getenv('REDIRECT_URI', default="http://127.0.0.1:8080"))
scope = config.get(section = 'Spotify', option = 'scope', fallback = 'playlist-modify-public, playlist-modify-private, user-read-email')
dnp_file = config.get(section = 'General', option = 'dnp_file', fallback = 'FIRST-Do-Not-Play-List-2025.xlsx')
remove_optional = config.getboolean(section = 'General', option = 'remove_optional', fallback = False)
remove_explicit = config.getboolean(section = 'General', option = 'remove_explicit', fallback = True)
make_public = config.getboolean(section = 'General', option = 'make_public', fallback = True)
enable_log = config.getboolean(section = 'General', option = 'enable_log', fallback = True)
log_file = config.get(section = 'General', option = 'log_file', fallback = 'cleaner.log')
log_level = config.get(section = 'General', option = 'log_level', fallback = os.getenv('LOG_LEVEL', default='INFO'))

match log_level:
    case 'DEBUG':
        real_log_level = logging.DEBUG
    case 'INFO':
        real_log_level = logging.INFO
    case 'WARNING':
        real_log_level = logging.WARNING
    case 'ERROR':
        real_log_level = logging.ERROR
    case 'CRITICAL':
        real_log_level = logging.CRITICAL
    case _:
        real_log_level = logging.INFO

if enable_log:
    logger = logging.getLogger()
    logging.basicConfig(filename = log_file, level = real_log_level)

    logger.info('## FRC/FTC Spotify Playlist Cleaner ##')
    logger.debug('Config: ')
    logger.debug('  [Spotify]')
    logger.debug('  redirect_uri = ' + redirect_uri)
    logger.debug('  scope = ' + scope)
    logger.debug('  [General]')
    logger.debug('  dnp_file = ' + dnp_file)
    logger.debug('  remove_optional = ' + str(remove_optional))
    logger.debug('  remove_explicit = ' + str(remove_explicit))
    logger.debug('  make_public = ' + str(make_public))
    logger.debug('  enable_log = ' + str(enable_log))
    logger.debug('  log_file = ' + log_file)
    logger.debug('  log_level = ' + log_level)

# create and configure the app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(64)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = './.flask-session/'
#app.config['APPLICATION_ROOT'] = os.getenv('APPLICATION_ROOT', default='/')
#app.config['PREFERRED_URL_SCHEME'] = os.getenv('SCHEME', default='http')
Session(app)

def _chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def _track_info(track_name, track_artist) -> str:
    return '\"' + track_name + '\" by \"' + track_artist + '\"'

def _log_playlist(playlist) -> None:
    logger.debug('Playlist Name: ' + playlist['name'])
    logger.debug('Playlist Description: ' + playlist['description'])
    if playlist['owner']['display_name'] is not None:
        logger.debug('Playlist Owner Name: ' + playlist['owner']['display_name'])
    logger.debug('Playlist Owner ID: ' + playlist['owner']['id'])
    logger.debug('Playlist Owner URI: ' + playlist['owner']['uri'])
    logger.debug('Playlist Owner URL: ' + playlist['owner']['external_urls']['spotify'])
    logger.debug('Playlist is Public: ' + str(playlist['public']))
    logger.debug('Playlist Size: ' + str(playlist['tracks']['total']))
    logger.debug('Playlist ID: ' + playlist['id'])
    logger.debug('Playlist URI: ' + playlist['uri'])
    logger.debug('Playlist URL: ' + playlist['external_urls']['spotify'])

def _search_for_track(track_name, track_artist, tracks) -> bool:
    skip = False
    logger.debug('Searching for Track \"' + track_name + '\" by \"' + track_artist + '\" in Do Not Play List.')
    matched_tracks = (index for index, track in enumerate(tracks) if track[0] == track_name)
    while True:
        track_index = next(matched_tracks, None)
        if track_index is None:
            logger.debug("Track Name does not exist in the Do Not Play List. Stopping Search.")
            break

        logger.debug('Track Name found in Do Not Play List, checking Artist.')
        logger.debug('Spotify Track Name: ' + track_name)
        logger.debug('DNP Track Name: ' + tracks[track_index][0])
        logger.debug('Track Artist: ' + track_artist)
        logger.debug('DNP Track Artist: ' + tracks[track_index][1])
        if track_artist == tracks[track_index][1]:
            logger.warning('Track ' + _track_info(track_name, track_artist) + ' found in Do Not Play List. Removing Track.')
            skip = True
            break
        logger.debug('Artist and Track combo not found in Do Not Play List. Searching the rest of the Do Not Play List')
    return skip

class LoadTracksFromWorkbookThread(threading.Thread):
    def __init__(self, file, worksheet = 'DO NOT PLAY LIST'):
        self.file = file
        self.worksheet = worksheet
        self.progress = 0.0
        self.tracks = []
        super().__init__()
    
    def run(self):
        max_row = self.file[self.worksheet].max_row

        count = 1
        for track in self.file[self.worksheet].iter_rows(min_col=2, max_col=3, min_row=2, max_row = max_row, values_only=True):
            if track != (None, None):
                self.tracks.append(track[::-1])
            count += 1
            self.progress = (float(count) / float(max_row)) * 100.0

class GetTracksFromPlaylistThread(threading.Thread):
    def __init__(self, playlist_id, playlist_total):
        self.playlist_id = playlist_id
        self.playlist_total = playlist_total
        self.progress = 0.0
        self.tracks = []
        super().__init__()

    def run(self):
        auth_manager = spotipy.oauth2.SpotifyOAuth(client_id=client_id,
                                                   client_secret=client_secret,
                                                   redirect_uri=redirect_uri,
                                                   scope=scope,
                                                   show_dialog=True)
        sp = spotipy.Spotify(auth_manager=auth_manager)

        logger.info('Getting tracks from Playlist.')

        offset = 0
        while True:
            logger.debug('Grabbing Tracks ' + str(offset + 1) + '-' + str(offset + min(100, self.playlist_total)) + '/' + str(self.playlist_total) + '.')
            response = sp.playlist_items(self.playlist_id,
                                        fields = 'items.track(artists,explicit,external_urls.spotify,id,name,uri),total',
                                        limit = 100,
                                        offset = offset,
                                        additional_types = ['track'])
            logger.debug('API reports Playlist contains ' + str(response['total']) + ' Tracks.')
            logger.info(str(len(response['items']) + offset) + '/' + str(self.playlist_total) + ' Tracks grabbed.')

            # If no items are returned, we have exhausted the playlist
            if len(response['items']) == 0:
                logger.debug('No Tracks returned, assuming Tracks have been exhausted.')
                break

            # Append the Track to the List.
            logger.debug('Adding Tracks to List.')
            self.tracks.extend(item['track'] for item in response['items'])
            logger.debug('List now contains ' + str(len(self.tracks)) + ' Tracks.')

            self.progress = (float(len(self.tracks)) / float(self.playlist_total)) * 100.0

            # Update the offset
            offset = offset + len(response['items'])
            if offset >= self.playlist_total:
                logger.debug("offset >= playlist_total. End of Playlist has been reached.")
                break

        logger.debug(str(len(self.tracks)) + ' Tracks out of ' + str(self.playlist_total) + ' expected Tracks grabbed.')
        logger.info(str(len(self.tracks)) + ' Tracks grabbed from Playlist.')

class GetCleanTracksFromTracksThread(threading.Thread):
    def __init__(self, tracks, dnp_tracks, optional_tracks):
        self.tracks = tracks
        self.dnp_tracks = dnp_tracks
        self.optional_tracks = optional_tracks
        self.progress = 0.0
        self.clean_tracks = []
        super().__init__()

    def run(self):
        logger.info('Cleaning ' + str(len(self.tracks)) + ' Tracks.')
        logger.debug(str(len(self.dnp_tracks)) + " Do Not Play Tracks loaded.")

        counter = 1
        for track in self.tracks:
            track_name = track['name']
            track_artist = track['artists'][0]['name']
            logger.debug('Current Track: ' + _track_info(track_name, track_artist))

            if track['explicit'] and remove_explicit:
                logger.warning('Track ' + _track_info(track_name, track_artist) + ' is Explicit and remove_explicit is set to True. Removing Track.')
                self.progress = (float(counter) / float(len(self.tracks))) * 100.0
                counter += 1
                continue

            skip = _search_for_track(track_name = track_name, track_artist = track_artist, tracks = self.dnp_tracks)
            if remove_optional and not skip:
                skip = _search_for_track(track_name = track_name, track_artist = track_artist, tracks = self.optional_tracks)

            if skip:
                self.progress = (float(counter) / float(len(self.tracks))) * 100.0
                counter += 1
                continue

            logger.info('Adding Track ' + _track_info(track_name, track_artist))
            self.clean_tracks.append(track)
            logger.info('List now contains ' + str(len(self.clean_tracks)) + ' Tracks.')
            self.progress = (float(counter) / float(len(self.tracks))) * 100.0

            counter += 1

        logger.info(str(len(self.clean_tracks)) + '/' + str(len(self.tracks)) + ' are Clean.')

class CreateCleanedPlaylistThread(threading.Thread):
    def __init__(self, user, unclean_playlist, playist_name, tracks):
        self.user = user
        self.unclean_playlist = unclean_playlist
        self.playist_name = playist_name
        self.tracks = tracks
        self.clean_playlist = None
        self.progress = 0.0
        super().__init__()

    def run(self):
        auth_manager = spotipy.oauth2.SpotifyOAuth(client_id=client_id,
                                                   client_secret=client_secret,
                                                   redirect_uri=redirect_uri,
                                                   scope=scope,
                                                   show_dialog=True)
        sp = spotipy.Spotify(auth_manager=auth_manager)

        logger.info('Creating Clean Playlist.')
        user_id = self.user['id']
        
        if self.playist_name == "":
            self.playist_name = self.unclean_playlist['name'] + " Cleaned"
        description = "Playlist cleaned for use in FRC/FTC Events by the FRC-FTC Spotify Playlist Cleaner. " + self.unclean_playlist['description']

        logger.debug('Calling Spotify API to create Playlist')
        self.clean_playlist = sp.user_playlist_create(user = user_id, name = self.playist_name, public = make_public, description = description)
        
        logger.info('Clean Playlist Created.')
        _log_playlist(self.clean_playlist)

        logger.info('Adding Clean Tracks to Clean Playlist')
        track_ids = [track['id'] for track in self.tracks]
        counter = 0
        for chunk in _chunks(track_ids, 100):
            logger.debug('Adding ' + str(len(chunk)) + ' Tracks.')
            sp.playlist_add_items(playlist_id = self.clean_playlist['id'], items = chunk)
            counter += len(chunk)
            logger.info(str(counter) + '/' + str(len(self.tracks)) + ' Tracks added to Playlist.')
            self.progress = (float(counter) / float(len(self.tracks))) * 100.0

class CleanPlaylistThread(threading.Thread):
    def __init__(self, user, url, clean_playlist_name):
        self.user = user
        self.url = url
        self.clean_playlist_name = clean_playlist_name
        self.clean_playlist = None
        self.progress = 0.0
        super().__init__()
    
    def run(self):
        auth_manager = spotipy.oauth2.SpotifyOAuth(client_id=client_id,
                                                   client_secret=client_secret,
                                                   redirect_uri=redirect_uri,
                                                   scope=scope,
                                                   show_dialog=True)
        sp = spotipy.Spotify(auth_manager=auth_manager)

        dnp = load_workbook(filename = dnp_file, read_only = True)

        # Load Do Not Play Tracks
        threads['dnp_tracks'] = LoadTracksFromWorkbookThread(dnp)
        threads['dnp_tracks'].start()
        threads['dnp_tracks'].join()
        dnp_tracks = threads['dnp_tracks'].tracks
        if remove_optional:
            self.progress += 20.0
        else:
            self.progress += 25.0

        # Load Optional Do Not Play Tracks
        optional_tracks = None
        if remove_optional:
            threads['optional_dnp_tracks'] = LoadTracksFromWorkbookThread(dnp, 'Optional')
            threads['optional_dnp_tracks'].start()
            threads['optional_dnp_tracks'].join()
            optional_tracks = threads['optional_dnp_tracks'].tracks
            self.progress += 20.0

        # Get Tracks from Playlist
        logger.debug('Grabbing Playlist from Spotify API')
        unclean_playlist = sp.playlist(self.url)
        logger.info('Spotify Playlist Grabbed')
        _log_playlist(unclean_playlist)
        threads['playlist_tracks'] = GetTracksFromPlaylistThread(playlist_id=self.url, playlist_total=unclean_playlist['tracks']['total'])
        threads['playlist_tracks'].start()
        threads['playlist_tracks'].join()
        playlist_items = threads['playlist_tracks'].tracks
        if remove_optional:
            self.progress += 20.0
        else:
            self.progress += 25.0

        # Get Clean Tracks
        threads['clean_tracks'] = GetCleanTracksFromTracksThread(tracks=playlist_items, dnp_tracks=dnp_tracks, optional_tracks=optional_tracks)
        threads['clean_tracks'].start()
        threads['clean_tracks'].join()
        clean_tracks = threads['clean_tracks'].clean_tracks
        if remove_optional:
            self.progress += 20.0
        else:
            self.progress += 25.0

        # Create Cleaned Playlist
        threads['cleaned_playlist'] = CreateCleanedPlaylistThread(user=self.user, unclean_playlist=unclean_playlist, playist_name=self.clean_playlist_name, tracks=clean_tracks)
        threads['cleaned_playlist'].start()
        threads['cleaned_playlist'].join()
        self.clean_playlist = threads['cleaned_playlist'].clean_playlist
        if remove_optional:
            self.progress += 20.0
        else:
            self.progress += 25.0

# ensure the instance folder exists
try:
    os.makedirs(app.instance_path)
except OSError:
    pass

@app.route('/', methods=('GET', 'POST'))
def index():
    global threads

    cache_handler = spotipy.cache_handler.FlaskSessionCacheHandler(session)
    auth_manager = spotipy.oauth2.SpotifyOAuth(client_id=client_id,
                                               client_secret=client_secret,
                                               redirect_uri=redirect_uri,
                                               scope=scope,
                                               cache_handler=cache_handler,
                                               show_dialog=True)
    sp = spotipy.Spotify(auth_manager=auth_manager)

    if request.args.get("code"):
        # Step 2. Being redirected from Spotify auth page
        auth_manager.get_access_token(request.args.get("code"))
        return redirect('/')

    if not auth_manager.validate_token(cache_handler.get_cached_token()):
        # Step 1. Display sign in link when no token
        auth_url = auth_manager.get_authorize_url()
        return f'<h2><a href="{auth_url}">Sign in</a></h2>'

    if request.method == "POST":
        user = sp.me()
        threads['clean_playlist'] = CleanPlaylistThread(user=user, url=request.form['playlist_url'], clean_playlist_name=request.form['clean_playlist_name'])
        threads['clean_playlist'].start()

        session.playlist_id = request.form['playlist_url'].rsplit('/', 1)[-1]
    return render_template('index.html')

@app.route('/sign_out')
def sign_out():
    session.pop("token_info", None)
    return redirect('/')

@app.route('/progress')
def progress():
    global threads

    progress = {}

    for thread_name, thread_object in threads.items():
        progress[thread_name] = threads[thread_name].progress

    return json.dumps(progress)

@app.route('/playlist_ids')
def playlist_ids():
    playlist_ids = {}

    if threads.get('clean_playlist') is not None:
        playlist_ids['clean_playlist_id'] = threads.get('clean_playlist').clean_playlist['id']
    else:
         playlist_ids['clean_playlist_id'] = 0

    return json.dumps(playlist_ids)

if __name__ == '__main__':
    app.run(threaded=True, port=int(os.getenv('PORT', default='8080')), debug=True)