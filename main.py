import configparser
import logging
import spotipy

from openpyxl import load_workbook
from spotipy.oauth2 import SpotifyOAuth

config = configparser.ConfigParser()
config.read('config')

client_id = config.get(section = 'Spotify', option = 'client_id', fallback = None)
client_secret = config.get(section = 'Spotify', option = 'client_secret', fallback = None)
redirect_uri = config.get(section = 'Spotify', option = 'redirect_uri', fallback = 'http://127.0.0.1:8080')
scope = config.get(section = 'Spotify', option = 'scope', fallback = 'playlist-modify-public, playlist-modify-private, user-read-email')
dnp_file = config.get(section = 'General', option = 'dnp_file', fallback = 'FIRST-Do-Not-Play-List-2025.xlsx')
remove_optional = config.getboolean(section = 'General', option = 'remove_optional', fallback = False)
remove_explicit = config.getboolean(section = 'General', option = 'remove_explicit', fallback = True)
make_public = config.getboolean(section = 'General', option = 'make_public', fallback = True)
enable_log = config.getboolean(section = 'General', option = 'enable_log', fallback = True)
log_file = config.get(section = 'General', option = 'log_file', fallback = 'cleaner.log')
log_level = config.get(section = 'General', option = 'log_level', fallback = 'INFO')

match log_level:
    case 'DEBUG':
        real_log_level = logging.DEBUG
    case 'INFO':
        real_log_level = logging.INFO
    case 'WARNING':
        real_log_level = logging.WARNING
    case 'ERROR':
        real_log_level = logging.ERROR
    case 'CRITICAL':
        real_log_level = logging.CRITICAL
    case _:
        real_log_level = logging.INFO

if enable_log:
    logger = logging.getLogger()
    logging.basicConfig(filename = log_file, level = real_log_level)

    logger.info('## FRC/FTC Spotify Playlist Cleaner ##')
    logger.debug('Config: ')
    logger.debug('  [Spotify]')
    logger.debug('  redirect_uri = ' + redirect_uri)
    logger.debug('  scope = ' + scope)
    logger.debug('  [General]')
    logger.debug('  dnp_file = ' + dnp_file)
    logger.debug('  remove_optional = ' + str(remove_optional))
    logger.debug('  remove_explicit = ' + str(remove_explicit))
    logger.debug('  make_public = ' + str(make_public))
    logger.debug('  enable_log = ' + str(enable_log))
    logger.debug('  log_file = ' + log_file)
    logger.debug('  log_level = ' + log_level)

def _chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def _track_info(track_name, track_artist) -> str:
    return '\"' + track_name + '\" by \"' + track_artist + '\"'

def _search_for_track(track_name, track_artist, tracks) -> bool:
    skip = False
    logger.debug('Searching for Track \"' + track_name + '\" by \"' + track_artist + '\" in Do Not Play List.')
    matched_tracks = (index for index, track in enumerate(tracks) if track[0] == track_name)
    while True:
        track_index = next(matched_tracks, None)
        if track_index is None:
            logger.debug("Track Name does not exist in the Do Not Play List. Stopping Search.")
            break

        logger.debug('Track Name found in Do Not Play List, checking Artist.')
        logger.debug('Spotify Track Name: ' + track_name)
        logger.debug('DNP Track Name: ' + tracks[track_index][0])
        logger.debug('Track Artist: ' + track_artist)
        logger.debug('DNP Track Artist: ' + tracks[track_index][1])
        if track_artist == tracks[track_index][1]:
            logger.warning('Track ' + _track_info(track_name, track_artist) + ' found in Do Not Play List. Removing Track.')
            skip = True
            break
        logger.debug('Artist and Track combo not found in Do Not Play List. Searching the rest of the Do Not Play List')
    return skip

def log_playlist(playlist) -> None:
    logger.debug('Playlist Name: ' + playlist['name'])
    logger.debug('Playlist Description: ' + playlist['description'])
    if playlist['owner']['display_name'] is not None:
        logger.debug('Playlist Owner Name: ' + playlist['owner']['display_name'])
    logger.debug('Playlist Owner ID: ' + playlist['owner']['id'])
    logger.debug('Playlist Owner URI: ' + playlist['owner']['uri'])
    logger.debug('Playlist Owner URL: ' + playlist['owner']['external_urls']['spotify'])
    logger.debug('Playlist is Public: ' + str(playlist['public']))
    logger.debug('Playlist Size: ' + str(playlist['tracks']['total']))
    logger.debug('Playlist ID: ' + playlist['id'])
    logger.debug('Playlist URI: ' + playlist['uri'])
    logger.debug('Playlist URL: ' + playlist['external_urls']['spotify'])

def load_tracks_from_workbook(file = None, worksheet = 'DO NOT PLAY LIST') -> list:
    tracks = []
    # Reverse tuple order to allow for searching by Song Name.
    for track in file[worksheet].iter_rows(min_col = 2, max_col = 3, min_row = 2, values_only = True):
        # Remove items that are empty
        if track != (None, None):
            tracks.append(track[::-1])
    return tracks

def get_tracks_from_playlist(sp = None, playlist_id = None, limit = 100, playlist_total = 0) -> list:
    logger.info('Getting tracks from Playlist.')

    offset = 0
    items = []
    while True:
        logger.debug('Grabbing Tracks ' + str(offset + 1) + '-' + str(offset + min(limit, playlist_total)) + '/' + str(playlist_total) + '.')
        response = sp.playlist_items(playlist_id,
                                     fields = 'items.track(artists,explicit,external_urls.spotify,id,name,uri),total',
                                     limit = limit,
                                     offset = offset,
                                     additional_types = ['track'])
        logger.debug('API reports Playlist contains ' + str(response['total']) + ' Tracks.')
        logger.info(str(len(response['items']) + offset) + '/' + str(playlist_total) + ' Tracks grabbed.')

        # If no items are returned, we have exhausted the playlist
        if len(response['items']) == 0:
            logger.debug('No Tracks returned, assuming Tracks have been exhausted.')
            break

        # Append the Track to the List.
        logger.debug('Adding Tracks to List.')
        items.extend(item['track'] for item in response['items'])
        logger.debug('List now contains ' + str(len(items)) + ' Tracks.')

        # Update the offset
        offset = offset + len(response['items'])

        if offset >= playlist_total:
            logger.debug("offset >= playlist_total. End of Playlist has been reached.")
            break

    logger.debug(str(len(items)) + ' Tracks out of ' + str(playlist_total) + ' expected Tracks grabbed.')
    logger.info(str(len(items)) + ' Tracks grabbed from Playlist.')
    return items

def get_clean_tracks_from_playlist(sp = None, tracks = None, dnp_tracks = None, optional_tracks = None) -> list:
    logger.info('Cleaning ' + str(len(tracks)) + ' Tracks.')
    logger.debug(str(len(dnp_tracks)) + " Do Not Play Tracks loaded.")

    items = []
    counter = 1
    for track in tracks:
        track_name = track['name']
        track_artist = track['artists'][0]['name']
        logger.debug('Current Track: ' + _track_info(track_name, track_artist))

        if track['explicit'] and remove_explicit:
            logger.warning('Track ' + _track_info(track_name, track_artist) + ' is Explicit and remove_explicit is set to True. Removing Track.')
            continue

        skip = _search_for_track(track_name = track_name, track_artist = track_artist, tracks = dnp_tracks)
        if remove_optional and not skip:
            skip = _search_for_track(track_name = track_name, track_artist = track_artist, tracks = optional_tracks)

        if skip:
            continue

        logger.info('Adding Track ' + _track_info(track_name, track_artist))
        items.append(track)
        logger.info('List now contains ' + str(len(items)) + ' Tracks.')
        counter += 1

    logger.info(str(len(items)) + '/' + str(len(tracks)) + ' are Clean.')
    return items

def create_cleaned_playlist(sp = None, user = None, unclean_playlist = None, name = "", tracks = None) -> dict:
    logger.info('Creating Clean Playlist.')
    user_id = user['id']
    
    if name == "":
        name = unclean_playlist['name'] + " Cleaned"
    description = "Playlist cleaned for use in FRC/FTC Events by the FRC-FTC Spotify Playlist Cleaner. " + unclean_playlist['description']

    logger.debug('Calling Spotify API to create Playlist')
    playlist = sp.user_playlist_create(user = user_id, name = name, public = make_public, description = description)
    
    logger.info('Clean Playlist Created.')
    _log_playlist(playlist)

    logger.info('Adding Clean Tracks to Clean Playlist')
    track_ids = [track['id'] for track in tracks]
    counter = 0
    for chunk in _chunks(track_ids, 100):
        logger.debug('Adding ' + str(len(chunk)) + ' Tracks.')
        sp.playlist_add_items(playlist_id = playlist['id'], items = chunk)
        counter += len(chunk)
        logger.info(str(counter) + '/' + str(len(tracks)) + ' Tracks added to Playlist.')
    return playlist

def main():
    logger.debug('Loading Do Not Play Workbook')
    dnp = load_workbook(filename = dnp_file,
                        read_only = True)
    logger.debug('Loaded Do Not Play Workbook')

    logger.debug('Initializing Spotify API')
    auth_manager = SpotifyOAuth(client_id = client_id,
                                client_secret = client_secret,
                                redirect_uri = redirect_uri,
                                scope = scope)
                    
    sp = spotipy.Spotify(auth_manager = auth_manager)
    user = sp.me()
    logger.info('Spotify API Connected as user ' + user['display_name'])
    logger.debug('Spotify User Email: ' + user['email'])
    logger.debug('Spotify User ID: ' + user['id'])
    logger.debug('Spotify User URI: ' + user['uri'])
    logger.debug('Spotify User URL: ' + user['external_urls']['spotify'])

    unclean_playlist_url = input('Enter the URL of the playlist to clean: ')
    logger.info('Spotify Playlist URL: ' + unclean_playlist_url)
    clean_playlist_name = input('Enter the name for the cleaned playlist (Enter to append \"FIRST Cleaned\" to the end of the original playlist name): ')
    logger.info('Cleaned Playlist Name: ' + clean_playlist_name)

    logger.debug('Grabbing Playlist from Spotify API')
    unclean_playlist = sp.playlist(unclean_playlist_url)
    logger.info('Spotify Playlist Grabbed')
    _log_playlist(unclean_playlist)

    dnp_tracks = load_tracks_from_workbook(file = dnp)
    optional_tracks = load_tracks_from_workbook(file = dnp, worksheet = 'Optional')
    playlist_items = get_tracks_from_playlist(sp = sp, playlist_id = unclean_playlist_url, playlist_total = unclean_playlist['tracks']['total'])
    clean_tracks = get_clean_tracks_from_playlist(sp = sp, tracks = playlist_items, dnp_tracks = dnp_tracks, optional_tracks = optional_tracks)
    clean_playlist = create_cleaned_playlist(sp = sp, user = user, unclean_playlist = unclean_playlist, name = clean_playlist_name, tracks = clean_tracks)